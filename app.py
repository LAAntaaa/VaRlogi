"""
╔══════════════════════════════════════════════════════════════╗
║        VaR ANALYTICS SUITE  ·  Streamlit App v3.0           ║
║        Département Gestion des Risques de Marché            ║
╚══════════════════════════════════════════════════════════════╝

Lancement :
    pip install streamlit yfinance pandas numpy scipy openpyxl reportlab matplotlib
    streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
from scipy.optimize import minimize
import io, os, warnings
warnings.filterwarnings("ignore")

# ── Imports optionnels ────────────────────────────────────────────────────────
try:
    import yfinance as yf
    HAS_YF = True
except ImportError:
    HAS_YF = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable, PageBreak, Image)
    from reportlab.lib.colors import HexColor
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG STREAMLIT + CSS
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="VaR Analytics Suite",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;600;700&family=JetBrains+Mono:wght@400;600&display=swap');

/* ── Base ─────────────────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Sora', sans-serif !important;
}
.stApp {
    background: linear-gradient(135deg, #0B1628 0%, #0d1e38 60%, #091422 100%);
}

/* ── Sidebar ──────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0a1525 0%, #0c1a30 100%) !important;
    border-right: 1px solid rgba(201,168,76,0.15) !important;
}
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: #8aa0bc !important;
    font-size: 12px !important;
}

/* ── Headers ──────────────────────────────────────────────────── */
h1 { 
    font-family: 'Sora', sans-serif !important;
    background: linear-gradient(90deg, #C9A84C, #e8c56a) !important;
    -webkit-background-clip: text !important;
    -webkit-text-fill-color: transparent !important;
    font-size: 1.9rem !important; font-weight: 700 !important;
    letter-spacing: 0.5px !important;
}
h2 { color: #d0daea !important; font-size: 1.15rem !important; font-weight: 600 !important; }
h3 { color: #C9A84C !important; font-size: 1rem !important; font-weight: 600 !important; }

/* ── Metric cards ─────────────────────────────────────────────── */
[data-testid="metric-container"] {
    background: rgba(17,31,53,0.85) !important;
    border: 1px solid rgba(201,168,76,0.18) !important;
    border-radius: 10px !important;
    padding: 14px 16px !important;
    box-shadow: 0 4px 20px rgba(0,0,0,0.3) !important;
    position: relative; overflow: hidden;
}
[data-testid="metric-container"]::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, #C9A84C, #2E6FD4);
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: #8aa0bc !important; font-size: 10px !important;
    text-transform: uppercase; letter-spacing: 1px;
    font-family: 'JetBrains Mono', monospace !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #C9A84C !important; font-size: 1.4rem !important;
    font-family: 'JetBrains Mono', monospace !important; font-weight: 700 !important;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"] {
    font-size: 10px !important;
    font-family: 'JetBrains Mono', monospace !important;
}

/* ── Dataframes / Tables ──────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(201,168,76,0.15) !important;
    border-radius: 8px !important;
    overflow: hidden !important;
}
.dvn-scroller { background: rgba(11,22,40,0.6) !important; }

/* ── Buttons ──────────────────────────────────────────────────── */
.stButton > button {
    background: linear-gradient(135deg, #1a3060, #243d78) !important;
    color: #C9A84C !important; border: 1px solid rgba(201,168,76,0.4) !important;
    border-radius: 8px !important; font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important; font-size: 13px !important;
    padding: 8px 22px !important; transition: all 0.2s !important;
    letter-spacing: 0.3px !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #243d78, #2E6FD4) !important;
    border-color: #C9A84C !important;
    box-shadow: 0 4px 16px rgba(201,168,76,0.25) !important;
    transform: translateY(-1px) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #C9A84C, #e8c56a) !important;
    color: #0B1628 !important; border-color: transparent !important;
}

/* ── Selectbox / Multiselect ──────────────────────────────────── */
.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: rgba(17,31,53,0.9) !important;
    border: 1px solid rgba(201,168,76,0.25) !important;
    border-radius: 8px !important; color: #d0daea !important;
}

/* ── Sliders ──────────────────────────────────────────────────── */
.stSlider > div > div > div > div {
    background: linear-gradient(90deg, #C9A84C, #2E6FD4) !important;
}

/* ── Info / Success / Warning / Error ────────────────────────── */
.stAlert { border-radius: 8px !important; border-left-width: 3px !important; }

/* ── Expander ─────────────────────────────────────────────────── */
.streamlit-expanderHeader {
    background: rgba(17,31,53,0.6) !important;
    border: 1px solid rgba(201,168,76,0.15) !important;
    border-radius: 8px !important; color: #d0daea !important;
}

/* ── Custom cards via st.markdown ────────────────────────────── */
.var-card {
    background: rgba(17,31,53,0.85);
    border: 1px solid rgba(201,168,76,0.18);
    border-radius: 10px; padding: 16px 20px;
    margin-bottom: 12px; position: relative; overflow: hidden;
}
.var-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, #C9A84C, #2E6FD4);
}
.var-card-title { font-size: 11px; color: #8aa0bc; text-transform: uppercase;
    letter-spacing: 1.5px; font-family: 'JetBrains Mono', monospace; margin-bottom: 6px; }
.var-card-value { font-size: 22px; font-weight: 700; color: #e05252;
    font-family: 'JetBrains Mono', monospace; }
.var-card-pct { font-size: 11px; color: #8aa0bc; font-family: 'JetBrains Mono', monospace; }
.var-card-es { font-size: 13px; color: #e87373; font-family: 'JetBrains Mono', monospace; margin-top: 4px; }
.badge-rec {
    display: inline-block; background: #C9A84C; color: #0B1628;
    font-size: 9px; font-weight: 700; padding: 2px 7px; border-radius: 4px;
    letter-spacing: 1px; font-family: 'JetBrains Mono', monospace; margin-left: 8px;
}
.section-header {
    border-left: 3px solid #C9A84C; padding-left: 12px;
    margin: 24px 0 12px; font-size: 14px; font-weight: 600; color: #d0daea;
    font-family: 'Sora', sans-serif;
}
.bt-ok { color: #1db87a; font-weight: 700; }
.bt-ko { color: #e05252; font-weight: 700; }
.info-box {
    background: rgba(46,111,212,0.08); border: 1px solid rgba(46,111,212,0.25);
    border-radius: 8px; padding: 14px 18px; margin: 12px 0;
    font-size: 12.5px; color: #c0d0e8; line-height: 1.7;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# DONNÉES & CACHE
# ══════════════════════════════════════════════════════════════════════════════

ACTIFS_DISPONIBLES = {
    "Apple (AAPL)":         "AAPL",
    "Microsoft (MSFT)":     "MSFT",
    "LVMH (MC.PA)":         "MC.PA",
    "TotalEnergies (TTE)":  "TTE.PA",
    "BNP Paribas (BNP)":    "BNP.PA",
    "Nestlé (NESN.SW)":     "NESN.SW",
    "SAP (SAP)":            "SAP",
    "Airbus (AIR.PA)":      "AIR.PA",
    "Tesla (TSLA)":         "TSLA",
    "Amazon (AMZN)":        "AMZN",
    "Nvidia (NVDA)":        "NVDA",
    "Safran (SAF.PA)":      "SAF.PA",
    "L'Oréal (OR.PA)":      "OR.PA",
    "ASML (ASML.AS)":       "ASML.AS",
    "Hermès (RMS.PA)":      "RMS.PA",
}

SECTEURS = {
    "AAPL": "Technologie", "MSFT": "Technologie", "MC.PA": "Luxe",
    "TTE.PA": "Énergie",   "BNP.PA": "Finance",   "NESN.SW": "Conso.",
    "SAP": "Technologie",  "AIR.PA": "Aéronaut.", "TSLA": "Auto/Tech",
    "AMZN": "Commerce",    "NVDA": "Technologie", "SAF.PA": "Aéronaut.",
    "OR.PA": "Beauté",     "ASML.AS": "Technologie", "RMS.PA": "Luxe",
}

@st.cache_data(show_spinner=False)
def telecharger_donnees(tickers: list, date_debut, date_fin) -> pd.DataFrame:
    if not HAS_YF:
        return pd.DataFrame()
    data = yf.download(tickers, start=str(date_debut), end=str(date_fin),
                       auto_adjust=True, progress=False)
    if isinstance(data.columns, pd.MultiIndex):
        if "Close" in data.columns.get_level_values(0):
            prix = data["Close"].copy()
        else:
            prix = data.xs(data.columns.get_level_values(0)[0], axis=1, level=0)
    else:
        if "Close" in data.columns:
            prix = data[["Close"]].copy()
            if len(tickers) == 1:
                prix.columns = tickers
        else:
            prix = data.copy()
    if isinstance(prix, pd.Series):
        prix = prix.to_frame()
    return prix.dropna(axis=1, how="all")


def donnees_simulation(tickers: list, n_days: int = 1500) -> pd.DataFrame:
    """Génère des prix simulés réalistes (GBM + corrélations) si yfinance indispo."""
    np.random.seed(42)
    n = len(tickers)
    mu_d  = np.full(n, 0.0004)
    sig_d = np.full(n, 0.012)
    corr  = np.eye(n) + 0.4 * (np.ones((n,n)) - np.eye(n))
    L = np.linalg.cholesky(corr)
    z = np.random.randn(n_days, n) @ L.T
    lr = mu_d + sig_d * z
    # Choc de marché
    mid = n_days // 2
    lr[mid:mid+20] *= 3.5
    prices = 100 * np.exp(np.cumsum(lr, axis=0))
    dates = pd.bdate_range(end="2024-12-31", periods=n_days)
    return pd.DataFrame(prices, index=dates, columns=tickers)


# ══════════════════════════════════════════════════════════════════════════════
# MOTEUR VAR — 7 MÉTHODES
# ══════════════════════════════════════════════════════════════════════════════

class VaREngine:
    def __init__(self, rendements: np.ndarray, pv: float = 10_000_000, horizon: int = 1):
        self.r       = rendements
        self.pv      = pv
        self.horizon = horizon
        self.n       = len(rendements)

    # ── 1. Historique ─────────────────────────────────────────────────────────
    def historique(self, alpha: float) -> dict:
        q   = np.percentile(self.r, (1-alpha)*100)
        exc = self.r[self.r <= q]
        es  = exc.mean() if len(exc) > 0 else q
        return {"VaR": -q*self.pv*np.sqrt(self.horizon),
                "ES":  -es*self.pv*np.sqrt(self.horizon),
                "VaR_pct": -q}

    # ── 2. Variance-Covariance ─────────────────────────────────────────────────
    def variance_covariance(self, alpha: float) -> dict:
        mu, sigma = self.r.mean(), self.r.std()
        z   = stats.norm.ppf(1-alpha)
        var = -(mu + z*sigma) * np.sqrt(self.horizon)
        es  = (sigma*stats.norm.pdf(z)/(1-alpha) - mu) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"μ={mu*100:.4f}%, σ={sigma*100:.4f}%"}

    # ── 3. RiskMetrics (EWMA λ=0.94) ──────────────────────────────────────────
    def riskmetrics(self, alpha: float, lam: float = 0.94) -> dict:
        r = self.r
        sig2 = np.zeros(len(r)); sig2[0] = r[0]**2
        for t in range(1, len(r)):
            sig2[t] = lam*sig2[t-1] + (1-lam)*r[t-1]**2
        sigma_t = np.sqrt(sig2[-1])
        z = stats.norm.ppf(1-alpha)
        var = -z * sigma_t * np.sqrt(self.horizon)
        es  = sigma_t * stats.norm.pdf(z) / (1-alpha) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"λ=0.94, σ_t={sigma_t*100:.4f}%"}

    # ── 4. Cornish-Fisher ─────────────────────────────────────────────────────
    def cornish_fisher(self, alpha: float) -> dict:
        mu, sigma = self.r.mean(), self.r.std()
        s = float(stats.skew(self.r))
        k = float(stats.kurtosis(self.r))
        z = stats.norm.ppf(1-alpha)
        z_cf = z + (z**2-1)*s/6 + (z**3-3*z)*k/24 - (2*z**3-5*z)*s**2/36
        var = -(mu + z_cf*sigma) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": var*self.pv*1.08, "VaR_pct": var,
                "params": f"z_CF={z_cf:.4f}, skew={s:.3f}, kurt={k:.3f}"}

    # ── 5. GARCH(1,1) ─────────────────────────────────────────────────────────
    def garch(self, alpha: float) -> dict:
        r = self.r
        def neg_ll(p):
            w, a, b = p
            if w<=0 or a<=0 or b<=0 or a+b>=1: return 1e10
            sig2 = np.zeros(len(r)); sig2[0] = np.var(r)
            for t in range(1,len(r)):
                sig2[t] = w + a*r[t-1]**2 + b*sig2[t-1]
            return 0.5*np.sum(np.log(2*np.pi*sig2) + r**2/sig2)
        try:
            res = minimize(neg_ll, [1e-6,0.08,0.89], method='L-BFGS-B',
                           bounds=[(1e-8,None),(0.001,0.3),(0.5,0.999)])
            w, a, b = res.x
        except Exception:
            w, a, b = 5e-7, 0.09, 0.90
        sig2 = np.zeros(len(r)); sig2[0] = np.var(r)
        for t in range(1,len(r)):
            sig2[t] = w + a*r[t-1]**2 + b*sig2[t-1]
        sigma_f = np.sqrt(w + a*r[-1]**2 + b*sig2[-1])
        z   = stats.norm.ppf(1-alpha)
        var = -z * sigma_f * np.sqrt(self.horizon)
        es  = sigma_f * stats.norm.pdf(z) / (1-alpha) * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": es*self.pv, "VaR_pct": var,
                "params": f"ω={w:.2e}, α={a:.3f}, β={b:.3f}"}

    # ── 6. TVE / POT ──────────────────────────────────────────────────────────
    def tve(self, alpha: float) -> dict:
        losses = -self.r
        u = np.percentile(losses, 90)
        exc = losses[losses > u] - u
        def neg_ll_gpd(p):
            xi, beta = p
            if beta <= 0: return 1e10
            u_ = exc / beta
            if xi != 0:
                if np.any(1+xi*u_ <= 0): return 1e10
                return len(u_)*np.log(beta) + (1+1/xi)*np.sum(np.log(1+xi*u_))
            return len(u_)*np.log(beta) + np.sum(u_)
        try:
            res = minimize(neg_ll_gpd, [0.1, np.std(exc)], method='L-BFGS-B',
                           bounds=[(-0.5,0.5),(1e-6,None)])
            xi, beta = res.x
        except Exception:
            xi, beta = 0.1, np.std(exc)
        n_u, n = len(exc), len(losses)
        p = 1 - alpha
        var = u + (beta/xi)*((n/n_u*p)**(-xi)-1) if xi != 0 \
              else u + beta*np.log(n/n_u*p)
        var = max(var, 0)
        es  = (var + beta - xi*u) / (1-xi)
        return {"VaR": var*self.pv*np.sqrt(self.horizon),
                "ES":  es*self.pv*np.sqrt(self.horizon),
                "VaR_pct": var,
                "params": f"ξ={xi:.3f}, β={beta:.4f}, u={u:.4f}"}

    # ── 7. TVE-GARCH ──────────────────────────────────────────────────────────
    def tve_garch(self, alpha: float) -> dict:
        r = self.r
        g = self.garch(alpha)
        w = float(g["params"].split(",")[0].split("=")[1])
        a_ = float(g["params"].split(",")[1].split("=")[1])
        b_ = float(g["params"].split(",")[2].split("=")[1])
        sig2 = np.zeros(len(r)); sig2[0] = np.var(r)
        for t in range(1,len(r)):
            sig2[t] = w + a_*r[t-1]**2 + b_*sig2[t-1]
        resid = r / np.sqrt(sig2)
        tve_res = VaREngine(resid, 1.0, 1).tve(alpha)
        sigma_f = np.sqrt(w + a_*r[-1]**2 + b_*sig2[-1])
        var = tve_res["VaR_pct"] * sigma_f * np.sqrt(self.horizon)
        return {"VaR": var*self.pv, "ES": var*self.pv*1.15, "VaR_pct": var,
                "params": f"GARCH+GPD hybride"}

    def compute_all(self, alphas=(0.95, 0.99)) -> dict:
        methods = {
            "Historique":          self.historique,
            "Variance-Covariance": self.variance_covariance,
            "RiskMetrics":         self.riskmetrics,
            "Cornish-Fisher":      self.cornish_fisher,
            "GARCH(1,1)":          self.garch,
            "TVE (POT)":           self.tve,
            "TVE-GARCH":           self.tve_garch,
        }
        results = {}
        for name, fn in methods.items():
            results[name] = {}
            for a in alphas:
                try:
                    results[name][a] = fn(a)
                except Exception as e:
                    results[name][a] = {"VaR": np.nan, "ES": np.nan,
                                        "VaR_pct": np.nan, "params": str(e)}
        return results


# ══════════════════════════════════════════════════════════════════════════════
# BACKTESTING — KUPIEC + CHRISTOFFERSEN
# ══════════════════════════════════════════════════════════════════════════════

def kupiec_test(rendements: np.ndarray, var_pct: float, alpha: float) -> dict:
    exc = (rendements < -var_pct).astype(int)
    N, T = int(exc.sum()), len(exc)
    if T == 0: return {"LR": np.nan, "p_value": np.nan, "valid": False, "N": 0, "T": 0}
    p0 = 1 - alpha
    p_hat = N / T
    if p_hat == 0:
        lr = -2 * T * np.log(1 - p0)
    elif p_hat == 1:
        lr = -2 * N * np.log(p0)
    else:
        lr = -2 * (T*np.log(1-p0) + N*np.log(p0)
                   - N*np.log(p_hat) - (T-N)*np.log(1-p_hat))
    pv = 1 - stats.chi2.cdf(max(lr,0), df=1)
    return {"LR": lr, "p_value": pv, "valid": pv > 0.05,
            "N": N, "T": T, "rate": p_hat, "expected": p0}


def christoffersen_test(rendements: np.ndarray, var_pct: float) -> dict:
    exc = (rendements < -var_pct).astype(int)
    n00 = np.sum((exc[:-1]==0) & (exc[1:]==0))
    n01 = np.sum((exc[:-1]==0) & (exc[1:]==1))
    n10 = np.sum((exc[:-1]==1) & (exc[1:]==0))
    n11 = np.sum((exc[:-1]==1) & (exc[1:]==1))
    pi01 = n01 / (n00+n01+1e-10)
    pi11 = n11 / (n10+n11+1e-10)
    pi   = (n01+n11) / (n00+n01+n10+n11+1e-10)
    try:
        lr = -2*(
            (n00+n10)*np.log(max(1-pi,1e-15))+(n01+n11)*np.log(max(pi,1e-15))
            - n00*np.log(max(1-pi01,1e-15)) - n01*np.log(max(pi01,1e-15))
            - n10*np.log(max(1-pi11,1e-15)) - n11*np.log(max(pi11,1e-15))
        )
    except Exception:
        lr = np.nan
    pv = 1 - stats.chi2.cdf(max(lr,0), df=1) if not np.isnan(lr) else np.nan
    return {"LR_ind": lr, "p_value_ind": pv,
            "valid": pv > 0.05 if not np.isnan(pv) else False}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS GRAPHIQUES (Matplotlib dark style)
# ══════════════════════════════════════════════════════════════════════════════

PLT_DARK = {
    "figure.facecolor": "#0d1e38", "axes.facecolor": "#0d1e38",
    "axes.edgecolor": "#2a3f5f", "axes.labelcolor": "#8aa0bc",
    "xtick.color": "#8aa0bc", "ytick.color": "#8aa0bc",
    "grid.color": "#1a2d48", "text.color": "#d0daea",
    "legend.facecolor": "#111f35", "legend.edgecolor": "#2a3f5f",
}

def fig_perf(rendements: pd.DataFrame, tickers: list) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, (ax1, ax2) = plt.subplots(2,1, figsize=(11,5),
                                        gridspec_kw={"height_ratios":[3,1]})
        colors = ["#2E6FD4","#C9A84C","#1db87a","#e05252","#a855f7","#f97316","#06b6d4","#ec4899"]
        port_r = rendements.mean(axis=1)
        cumul  = (1+port_r).cumprod()
        ax1.fill_between(range(len(cumul)), cumul, 1, alpha=0.1, color="#2E6FD4")
        ax1.plot(cumul.values, color="#2E6FD4", lw=1.8, label="Portefeuille")
        for i,t in enumerate(tickers[:4]):
            if t in rendements.columns:
                c = (1+rendements[t]).cumprod()
                ax1.plot(c.values, color=colors[(i+1)%len(colors)], lw=0.8, alpha=0.5, label=t)
        ax1.axhline(1, color="#C9A84C", lw=0.7, ls="--", alpha=0.6)
        ax1.set_title("Performance cumulée", fontsize=10, color="#C9A84C", pad=6)
        ax1.legend(fontsize=7, loc="upper left"); ax1.grid(True, alpha=0.3)
        ax1.tick_params(labelsize=8)
        col_bars = ["#1db87a" if v>=0 else "#e05252" for v in port_r]
        ax2.bar(range(len(port_r)), port_r*100, color=col_bars, alpha=0.7, width=1)
        ax2.axhline(0, color="#C9A84C", lw=0.5)
        ax2.set_title("Rendements journaliers (%)", fontsize=9, color="#8aa0bc", pad=4)
        ax2.tick_params(labelsize=7)
        ax2.grid(True, alpha=0.2)
        plt.tight_layout(); return fig


def fig_var_comparaison(var_results: dict, conf: float, pv: float) -> plt.Figure:
    methods = list(var_results.keys())
    vars_   = [var_results[m][conf]["VaR"]/1000 for m in methods]
    ess_    = [var_results[m][conf]["ES"]/1000  for m in methods]
    colors  = ["#2E6FD4","#C9A84C","#1db87a","#e05252","#a855f7","#f97316","#06b6d4"]
    short   = [m.replace("Variance-Covariance","VCV").replace("Cornish-Fisher","C-Fisher")
               .replace("RiskMetrics","RiskM.") for m in methods]
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(11,4))
        x = np.arange(len(methods)); w = 0.38
        b1 = ax.bar(x-w/2, vars_, w, color=colors, alpha=0.85, label="VaR")
        b2 = ax.bar(x+w/2, ess_,  w, color=colors, alpha=0.45, label="ES")
        for b in b1:
            ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.5,
                    f"{b.get_height():.0f}k", ha="center", va="bottom",
                    fontsize=7, color="#e05252", fontweight="bold")
        ax.set_xticks(x); ax.set_xticklabels(short, rotation=25, ha="right", fontsize=8)
        ax.set_ylabel("k€", fontsize=9); ax.grid(axis="y", alpha=0.25)
        ax.set_title(f"VaR & ES — Niveau de confiance {conf*100:.0f}% — Portefeuille {pv/1e6:.0f}M€",
                     fontsize=10, color="#C9A84C", pad=8)
        ax.legend(fontsize=8); plt.tight_layout(); return fig


def fig_distribution(rendements: np.ndarray, var_results: dict) -> plt.Figure:
    with plt.rc_context(PLT_DARK):
        fig, ax = plt.subplots(figsize=(11,4))
        ax.hist(rendements*100, bins=70, density=True, color="#2E6FD4",
                alpha=0.55, edgecolor="#1a3060", lw=0.3, label="Rendements obs.")
        mu, sig = rendements.mean(), rendements.std()
        x = np.linspace(rendements.min(), rendements.max(), 300)
        ax.plot(x*100, stats.norm.pdf(x,mu,sig)/100, color="#C9A84C",
                lw=2, ls="--", label="N(μ,σ)")
        colors_v = {"Historique":"#e05252","GARCH(1,1)":"#a855f7","TVE (POT)":"#f97316"}
        for meth, col in colors_v.items():
            if meth in var_results:
                p = var_results[meth].get(0.99,{}).get("VaR_pct")
                if p and not np.isnan(p):
                    ax.axvline(-p*100, color=col, lw=1.5, ls=":",
                               label=f"VaR 99% {meth}")
        ax.set_xlabel("Rendement journalier (%)", fontsize=9)
        ax.set_ylabel("Densité", fontsize=9)
        ax.set_title("Distribution des rendements & VaR 99%", fontsize=10,
                     color="#C9A84C", pad=8)
        ax.legend(fontsize=7); ax.grid(True, alpha=0.2)
        plt.tight_layout(); return fig


def fig_backtesting(bt_results: dict) -> plt.Figure:
    methods = list(bt_results.keys())
    short   = [m.replace("Variance-Covariance","VCV").replace("Cornish-Fisher","C-Fisher")
               .replace("RiskMetrics","RiskM.") for m in methods]
    p95 = [bt_results[m][0.95]["p_value"] for m in methods]
    p99 = [bt_results[m][0.99]["p_value"] for m in methods]
    with plt.rc_context(PLT_DARK):
        fig, axes = plt.subplots(1,2, figsize=(11,3.5))
        for ax, pvals, title in zip(axes, [p95, p99], ["Kupiec 95%","Kupiec 99%"]):
            cols = ["#1db87a" if p>0.05 else "#e05252" for p in pvals]
            ax.bar(short, pvals, color=cols, alpha=0.85)
            ax.axhline(0.05, color="#C9A84C", lw=1.8, ls="--", label="Seuil 5%")
            ax.set_xticks(range(len(short)))
            ax.set_xticklabels(short, rotation=30, ha="right", fontsize=7.5)
            ax.set_ylabel("p-value", fontsize=9); ax.grid(axis="y", alpha=0.2)
            ax.set_title(title, fontsize=10, color="#C9A84C", pad=6)
            ax.legend(fontsize=8)
        plt.tight_layout(); return fig


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generer_excel(rendements_df: pd.DataFrame, var_results: dict,
                   bt_results: dict, pv: float) -> bytes | None:
    if not HAS_XLSX:
        return None
    wb = Workbook()

    NAVY, BLUE, GOLD = "1B2A4A", "2E5FA3", "C9A84C"
    WHITE, LGRAY = "FFFFFF", "F0F4FA"

    def th(ws, r, c, v, bg=NAVY, fg=WHITE):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Calibri", bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        s = Side(border_style="thin", color="CCCCCC")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def td(ws, r, c, v, bg=WHITE):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        s = Side(border_style="thin", color="DDDDDD")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    # ── Feuille 1 : Résumé ────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "📊 Résumé VaR"
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]; c.value = "RAPPORT VaR — SYNTHÈSE DES RÉSULTATS"
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    headers = ["Méthode","VaR 95% (€)","VaR 95% (%)","VaR 99% (€)","VaR 99% (%)","ES 99% (€)"]
    for j,h in enumerate(headers,1): th(ws1, 2, j, h, bg=BLUE)

    for i,(m,res) in enumerate(var_results.items(), 3):
        r95 = res.get(0.95,{}); r99 = res.get(0.99,{})
        bg = LGRAY if i%2==0 else WHITE
        td(ws1,i,1,m,bg=bg)
        td(ws1,i,2,round(r95.get("VaR",0),0),bg=bg)
        td(ws1,i,3,f"{r95.get('VaR_pct',0)*100:.3f}%",bg=bg)
        td(ws1,i,4,round(r99.get("VaR",0),0),bg=bg)
        td(ws1,i,5,f"{r99.get('VaR_pct',0)*100:.3f}%",bg=bg)
        td(ws1,i,6,round(r99.get("ES",0),0),bg=bg)

    for w,col in zip([26,16,12,16,12,16],["A","B","C","D","E","F"]):
        ws1.column_dimensions[col].width = w

    # ── Feuille 2 : Backtesting ────────────────────────────────────────────────
    ws2 = wb.create_sheet("🧪 Backtesting")
    hdrs = ["Méthode","CL","Exceptions","T","Taux obs.","Taux att.",
            "Kupiec LR","Kupiec p","Kupiec OK","CC LR","CC p","CC OK"]
    for j,h in enumerate(hdrs,1): th(ws2,1,j,h,bg=BLUE)

    row = 2
    for m, alphas in bt_results.items():
        for a, res in alphas.items():
            bg = LGRAY if row%2==0 else WHITE
            k = res.get("kupiec",{}); cc = res.get("cc",{})
            td(ws2,row,1,m,bg=bg); td(ws2,row,2,f"{a*100:.0f}%",bg=bg)
            td(ws2,row,3,k.get("N",""),bg=bg); td(ws2,row,4,k.get("T",""),bg=bg)
            td(ws2,row,5,f"{k.get('rate',0)*100:.2f}%",bg=bg)
            td(ws2,row,6,f"{(1-a)*100:.2f}%",bg=bg)
            td(ws2,row,7,round(k.get("LR",0),3),bg=bg)
            td(ws2,row,8,round(k.get("p_value",0),4),bg=bg)
            td(ws2,row,9,"OUI" if k.get("valid") else "NON",bg=bg)
            td(ws2,row,10,round(cc.get("LR_ind",0),3),bg=bg)
            td(ws2,row,11,round(cc.get("p_value_ind",0),4),bg=bg)
            td(ws2,row,12,"OUI" if cc.get("valid") else "NON",bg=bg)
            row += 1
    for w,col in zip([26,6,12,8,10,10,10,10,10,10,10,10],
                     ["A","B","C","D","E","F","G","H","I","J","K","L"]):
        ws2.column_dimensions[col].width = w

    # ── Feuille 3 : Rendements ────────────────────────────────────────────────
    ws3 = wb.create_sheet("📋 Données")
    r_port = rendements_df.mean(axis=1).tail(250)
    for j,h in enumerate(["Date","Rdt Portfolio (%)"],1): th(ws3,1,j,h,bg=BLUE)
    for i,(d,v) in enumerate(r_port.items(),2):
        bg = LGRAY if i%2==0 else WHITE
        td(ws3,i,1,d.strftime("%d/%m/%Y"),bg=bg)
        td(ws3,i,2,round(v*100,4),bg=bg)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT PDF
# ══════════════════════════════════════════════════════════════════════════════

def generer_pdf(var_results: dict, bt_results: dict, metrics: dict, pv: float,
                fig_var=None, fig_dist=None) -> bytes | None:
    if not HAS_PDF:
        return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    NAVY_C  = HexColor("#1B2A4A"); BLUE_C = HexColor("#2E5FA3")
    GOLD_C  = HexColor("#C9A84C"); RED_C  = HexColor("#C0392B")
    GREEN_C = HexColor("#1F7A4D"); LGRAY_C= HexColor("#F0F4FA")

    S = lambda name, **kw: ParagraphStyle(name, **kw)
    s_title = S("t", fontName="Helvetica-Bold", fontSize=22, textColor=HexColor("#FFFFFF"),
                alignment=TA_CENTER, spaceAfter=4)
    s_sub   = S("s", fontName="Helvetica-Oblique", fontSize=11, textColor=GOLD_C,
                alignment=TA_CENTER, spaceAfter=6)
    s_h2    = S("h2", fontName="Helvetica-Bold", fontSize=12, textColor=NAVY_C,
                spaceBefore=12, spaceAfter=6)
    s_body  = S("b", fontName="Helvetica", fontSize=9, textColor=HexColor("#3A3A3A"),
                alignment=TA_JUSTIFY, spaceAfter=6, leading=14)

    def tbl_style():
        return TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BLUE_C),("TEXTCOLOR",(0,0),(-1,0),HexColor("#FFFFFF")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[HexColor("#FFFFFF"),LGRAY_C]),
            ("GRID",(0,0),(-1,-1),0.3,HexColor("#CCCCCC")),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ])

    story = []
    # Couverture
    story.append(Spacer(1,2*cm))
    cover = Table([[Paragraph("RAPPORT DE GESTION DES RISQUES", s_title)]], [15*cm])
    cover.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),NAVY_C),
                                ("TOPPADDING",(0,0),(-1,-1),18),("BOTTOMPADDING",(0,0),(-1,-1),18)]))
    story.append(cover)
    story.append(Spacer(1,0.3*cm))
    cover2 = Table([[Paragraph("Value at Risk — 7 méthodes · Backtesting · Analyse complète", s_sub)]],[15*cm])
    cover2.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),HexColor("#243B60")),
                                  ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    story.append(cover2)
    story.append(Spacer(1,1.5*cm))

    from datetime import date
    info = [["Valeur portefeuille", f"{pv:,.0f} €"],
            ["Horizon", "1 jour ouvrée"],
            ["Niveaux de confiance", "95% et 99%"],
            ["Date de production", date.today().strftime("%d/%m/%Y")]]
    t_info = Table([[Paragraph(k,s_body),Paragraph(v,s_body)] for k,v in info], [5*cm,10*cm])
    t_info.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
                                  ("ROWBACKGROUNDS",(0,0),(-1,-1),[HexColor("#FFFFFF"),LGRAY_C]),
                                  ("GRID",(0,0),(-1,-1),0.3,HexColor("#DDDDDD")),
                                  ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(t_info)
    story.append(PageBreak())

    # Résultats VaR
    story.append(Paragraph("1. RÉSULTATS DE LA VALUE AT RISK", s_h2))
    story.append(HRFlowable(width="100%",thickness=1,color=GOLD_C,spaceAfter=8))
    var_data = [["Méthode","VaR 95% (€)","VaR 95% (%)","VaR 99% (€)","VaR 99% (%)","ES 99% (€)"]]
    for m, res in var_results.items():
        r95, r99 = res.get(0.95,{}), res.get(0.99,{})
        var_data.append([m,
            f"{r95.get('VaR',0):,.0f} €", f"{r95.get('VaR_pct',0)*100:.3f}%",
            f"{r99.get('VaR',0):,.0f} €", f"{r99.get('VaR_pct',0)*100:.3f}%",
            f"{r99.get('ES',0):,.0f} €"])
    t_var = Table(var_data, [3.5*cm,2.5*cm,2*cm,2.5*cm,2*cm,2.5*cm])
    t_var.setStyle(tbl_style())
    story.append(t_var)
    story.append(Spacer(1,0.5*cm))

    # Graphique VaR
    if fig_var:
        buf_img = io.BytesIO(); fig_var.savefig(buf_img, format="png", dpi=120, bbox_inches="tight"); buf_img.seek(0)
        story.append(Image(buf_img, width=15*cm, height=5*cm))
        plt.close(fig_var)
    story.append(PageBreak())

    # Backtesting
    story.append(Paragraph("2. BACKTESTING", s_h2))
    story.append(HRFlowable(width="100%",thickness=1,color=GOLD_C,spaceAfter=8))
    story.append(Paragraph(
        "Test de Kupiec (POF) : H₀ → fréquence observée = 1−α. "
        "Test de Christoffersen : teste l'indépendance temporelle des exceptions. "
        "p > 0.05 → modèle non rejeté.", s_body))
    bt_data = [["Méthode","CL","Exceptions","Taux obs.","Kupiec p","Kupiec","CC p","CC"]]
    for m, alphas in bt_results.items():
        for a, res in alphas.items():
            k = res.get("kupiec",{}); cc = res.get("cc",{})
            bt_data.append([m, f"{a*100:.0f}%", str(k.get("N","")),
                f"{k.get('rate',0)*100:.2f}%",
                f"{k.get('p_value',0):.4f}", "✓ OUI" if k.get("valid") else "✗ NON",
                f"{cc.get('p_value_ind',0):.4f}", "✓ OUI" if cc.get("valid") else "✗ NON"])
    t_bt = Table(bt_data, [3.5*cm,1.2*cm,1.5*cm,1.8*cm,1.8*cm,1.5*cm,1.8*cm,1.5*cm])
    t_bt.setStyle(tbl_style())
    story.append(t_bt)
    story.append(Spacer(1,0.5*cm))

    if fig_dist:
        buf_img2 = io.BytesIO(); fig_dist.savefig(buf_img2, format="png", dpi=120, bbox_inches="tight"); buf_img2.seek(0)
        story.append(Image(buf_img2, width=15*cm, height=5*cm))
        plt.close(fig_dist)

    doc.build(story)
    buf.seek(0); return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════

for key in ["prix","rendements","var_results","bt_results","actifs_choisis","pv"]:
    if key not in st.session_state:
        st.session_state[key] = None


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 📉 VaR Analytics Suite")
    st.markdown("<div style='font-size:10px;color:#8aa0bc;font-family:JetBrains Mono,monospace;margin-bottom:16px'>v3.0 · Département Risque</div>", unsafe_allow_html=True)
    st.divider()

    menu = st.selectbox(
        "Navigation",
        ["🏠 Accueil", "🏦 Portefeuille", "📉 Calcul VaR",
         "🧪 Backtesting", "📊 Reporting"],
        label_visibility="collapsed"
    )
    st.divider()

    # Mini résumé portefeuille si données chargées
    if st.session_state["rendements"] is not None:
        r = st.session_state["rendements"].mean(axis=1)
        ann_r = r.mean() * 252
        ann_v = r.std() * np.sqrt(252)
        sharpe = (r.mean() - 0.03/252) / r.std() * np.sqrt(252)
        st.markdown("**Portefeuille chargé**")
        st.markdown(f"""
        <div style='font-size:11px;font-family:JetBrains Mono,monospace;line-height:1.9'>
        <span style='color:#8aa0bc'>Rdt ann. :</span> <span style='color:#1db87a'>+{ann_r*100:.2f}%</span><br>
        <span style='color:#8aa0bc'>Vol. ann. :</span> <span style='color:#d0daea'>{ann_v*100:.2f}%</span><br>
        <span style='color:#8aa0bc'>Sharpe   :</span> <span style='color:#C9A84C'>{sharpe:.3f}</span>
        </div>""", unsafe_allow_html=True)
        st.divider()

    st.markdown("""
    <div style='font-size:10px;color:#8aa0bc;line-height:1.8'>
    <b style='color:#C9A84C'>7 méthodes</b><br>
    · Historique<br>· Variance-Covariance<br>· RiskMetrics<br>
    · Cornish-Fisher<br>· GARCH(1,1)<br>· TVE (POT)<br>· TVE-GARCH
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : ACCUEIL
# ══════════════════════════════════════════════════════════════════════════════

if menu == "🏠 Accueil":
    st.title("VaR Analytics Suite")
    st.markdown("""
    <div class='info-box'>
    Progiciel professionnel de calcul, comparaison et validation de la <b>Value at Risk</b>
    sur un portefeuille d'actions. Développé selon les standards <b>Bâle III/IV</b>.
    </div>""", unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Méthodes VaR", "7", "Complètes")
    with col2: st.metric("Tests backtest", "2", "Kupiec + CC")
    with col3: st.metric("Actifs dispo.", len(ACTIFS_DISPONIBLES), "Euros & US")
    with col4: st.metric("Export", "Excel + PDF", "Auto")

    st.markdown("<div class='section-header'>Fonctionnalités</div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        **📥 Données de marché**
        - Téléchargement automatique Yahoo Finance
        - 15 actifs prédéfinis + saisie libre
        - Simulation intégrée si hors ligne

        **📉 Calcul des VaR**
        - 7 méthodes simultanées
        - Niveaux 95% et 99%
        - Expected Shortfall (ES/CVaR)
        """)
    with c2:
        st.markdown("""
        **🧪 Backtesting**
        - Test de Kupiec (POF) — fréquence
        - Test de Christoffersen — indépendance
        - p-values et verdict automatique

        **📊 Reporting**
        - Export Excel 3 feuilles formaté
        - Rapport PDF exécutif avec graphiques
        - Prêt à envoyer à la Direction Risque
        """)

    st.markdown("<div class='section-header'>Équipe Projet</div>", unsafe_allow_html=True)
    cols = st.columns(4)
    membres = ["Anta Mbaye", "Harlem D. Adjagba", "Ecclésiaste Gnargo", "Wariol G. Kopangoye"]
    for col, m in zip(cols, membres):
        with col:
            st.markdown(f"""
            <div class='var-card' style='text-align:center;padding:14px'>
            <div style='font-size:22px;margin-bottom:6px'>👤</div>
            <div style='font-size:11px;font-weight:600;color:#d0daea'>{m}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style='text-align:center;margin-top:16px;font-size:11px;color:#8aa0bc'>
    Double diplôme M2 IFIM · Ing 3 MACS — Mathématiques Appliquées au Calcul Scientifique
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : PORTEFEUILLE
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "🏦 Portefeuille":
    st.title("Construction du Portefeuille")

    st.markdown("<div class='section-header'>Sélection des actifs</div>", unsafe_allow_html=True)

    col1, col2 = st.columns([2,1])
    with col1:
        actifs_choisis = st.multiselect(
            "Actifs financiers",
            list(ACTIFS_DISPONIBLES.keys()),
            default=["Apple (AAPL)", "Microsoft (MSFT)", "LVMH (MC.PA)",
                     "TotalEnergies (TTE)", "BNP Paribas (BNP)"],
            help="Sélectionnez entre 2 et 10 actifs"
        )
    with col2:
        pv_millions = st.number_input(
            "Valeur du portefeuille (M€)",
            min_value=0.1, max_value=1000.0, value=10.0, step=0.5
        )
        pv = pv_millions * 1_000_000

    col3, col4, col5 = st.columns(3)
    with col3:
        date_debut = st.date_input("Date de début", value=pd.to_datetime("2019-01-01"))
    with col4:
        date_fin = st.date_input("Date de fin", value=pd.to_datetime("today"))
    with col5:
        source = st.radio("Source données", ["Yahoo Finance", "Simulation"],
                          horizontal=True,
                          help="Simulation : données synthétiques réalistes si pas de connexion")

    btn_col, _ = st.columns([1,3])
    with btn_col:
        btn = st.button("▶  Charger les données", type="primary", use_container_width=True)

    if btn:
        if len(actifs_choisis) < 2:
            st.warning("Sélectionnez au moins 2 actifs.")
        elif date_debut >= date_fin:
            st.warning("La date de début doit être antérieure à la date de fin.")
        else:
            tickers = [ACTIFS_DISPONIBLES[a] for a in actifs_choisis]
            with st.spinner("Chargement des données en cours…"):
                if source == "Yahoo Finance" and HAS_YF:
                    prix = telecharger_donnees(tickers, date_debut, date_fin)
                    if prix.empty:
                        st.warning("Aucune donnée récupérée, passage en simulation.")
                        prix = donnees_simulation(tickers)
                else:
                    prix = donnees_simulation(tickers)

                rendements = prix.pct_change(fill_method=None).dropna(how="all")
                rendements = rendements.dropna(axis=1, how="any")
                prix = prix[rendements.columns]
                tickers_valides = list(rendements.columns)

            st.session_state["prix"] = prix
            st.session_state["rendements"] = rendements
            st.session_state["actifs_choisis"] = tickers_valides
            st.session_state["pv"] = pv
            st.session_state["var_results"] = None
            st.session_state["bt_results"]  = None
            st.success(f"✅ {len(tickers_valides)} actif(s) chargés — {len(rendements)} jours de données.")

    # Affichage si données disponibles
    if st.session_state["rendements"] is not None:
        rendements = st.session_state["rendements"]
        prix       = st.session_state["prix"]

        st.markdown("<div class='section-header'>Statistiques du portefeuille</div>", unsafe_allow_html=True)

        port_r = rendements.mean(axis=1)
        ann_r  = port_r.mean() * 252
        ann_v  = port_r.std()  * np.sqrt(252)
        sharpe = (port_r.mean() - 0.03/252) / port_r.std() * np.sqrt(252)
        skew   = float(stats.skew(port_r))
        kurt   = float(stats.kurtosis(port_r))
        mdd    = float(((1+port_r).cumprod() / (1+port_r).cumprod().cummax() - 1).min())

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Rdt Annualisé", f"+{ann_r*100:.2f}%")
        c2.metric("Vol. Annuelle",  f"{ann_v*100:.2f}%")
        c3.metric("Sharpe",         f"{sharpe:.3f}")
        c4.metric("Max Drawdown",   f"{mdd*100:.2f}%")
        c5.metric("Skewness",       f"{skew:.4f}")
        c6.metric("Kurtosis (exc)", f"{kurt:.4f}")

        st.markdown("<div class='section-header'>Performance & Rendements</div>", unsafe_allow_html=True)
        fig = fig_perf(rendements, list(rendements.columns))
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

        st.markdown("<div class='section-header'>Statistiques individuelles</div>", unsafe_allow_html=True)
        stats_df = pd.DataFrame({
            "Ticker":       rendements.columns,
            "Secteur":      [SECTEURS.get(t,"—") for t in rendements.columns],
            "Rdt moy. (%)": (rendements.mean()*252*100).round(2),
            "Vol. ann. (%)": (rendements.std()*np.sqrt(252)*100).round(2),
            "Skewness":     rendements.apply(lambda c: round(stats.skew(c),4)),
            "Kurtosis":     rendements.apply(lambda c: round(stats.kurtosis(c),4)),
            "Min (%)":      (rendements.min()*100).round(3),
            "Max (%)":      (rendements.max()*100).round(3),
        }).set_index("Ticker")
        st.dataframe(stats_df, use_container_width=True)

        with st.expander("📋 Données brutes (50 derniers jours)"):
            st.dataframe(prix.tail(50), use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : VAR
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "📉 Calcul VaR":
    st.title("Calcul de la Value at Risk")

    if st.session_state["rendements"] is None:
        st.info("💡 Commencez par charger un portefeuille dans la page **Portefeuille**.")
        st.stop()

    rendements = st.session_state["rendements"]
    pv         = st.session_state["pv"] or 10_000_000
    port_r     = rendements.mean(axis=1).values

    st.markdown("<div class='section-header'>Paramètres de calcul</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        horizon = st.slider("Horizon (jours)", 1, 10, 1)
    with c2:
        conf_options = st.multiselect("Niveaux de confiance",
                                       [0.90, 0.95, 0.975, 0.99],
                                       default=[0.95, 0.99],
                                       format_func=lambda x: f"{x*100:.1f}%")
    with c3:
        methodes_sel = st.multiselect("Méthodes à calculer",
                                       ["Historique","Variance-Covariance","RiskMetrics",
                                        "Cornish-Fisher","GARCH(1,1)","TVE (POT)","TVE-GARCH"],
                                       default=["Historique","Variance-Covariance",
                                                "RiskMetrics","Cornish-Fisher",
                                                "GARCH(1,1)","TVE (POT)","TVE-GARCH"])

    btn2, _ = st.columns([1,3])
    with btn2:
        calc_btn = st.button("▶  Calculer les 7 VaR", type="primary", use_container_width=True)

    if calc_btn:
        if not conf_options:
            st.warning("Sélectionnez au moins un niveau de confiance.")
        else:
            with st.spinner("Calcul en cours…"):
                engine = VaREngine(port_r, pv, horizon)
                var_results = engine.compute_all(tuple(sorted(conf_options)))
                # Filtrer les méthodes non sélectionnées
                var_results = {k:v for k,v in var_results.items() if k in methodes_sel}
                st.session_state["var_results"] = var_results
            st.success(f"✅ VaR calculée pour {len(var_results)} méthodes × {len(conf_options)} niveaux de confiance.")

    if st.session_state["var_results"]:
        var_results = st.session_state["var_results"]
        alphas_used = sorted(list(list(var_results.values())[0].keys()))
        alpha_display = st.select_slider("Afficher pour :",
                                          options=alphas_used,
                                          format_func=lambda x: f"{x*100:.0f}%",
                                          value=alphas_used[-1])

        st.markdown("<div class='section-header'>Résultats par méthode</div>", unsafe_allow_html=True)

        METHODE_RECOMMANDEE = "TVE-GARCH"
        cols = st.columns(min(len(var_results), 4))
        for i, (method, res) in enumerate(var_results.items()):
            r = res.get(alpha_display, {})
            var_val = r.get("VaR", np.nan)
            pct_val = r.get("VaR_pct", np.nan)
            es_val  = r.get("ES",  np.nan)
            rec = f'<span class="badge-rec">★ Recommandé</span>' if method == METHODE_RECOMMANDEE else ""
            with cols[i % len(cols)]:
                st.markdown(f"""
                <div class='var-card'>
                  <div class='var-card-title'>{method}{rec}</div>
                  <div class='var-card-value'>{var_val/1000:.1f} k€</div>
                  <div class='var-card-pct'>VaR {alpha_display*100:.0f}% · {pct_val*100:.3f}%</div>
                  <div class='var-card-es'>ES : {es_val/1000:.1f} k€</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-header'>Tableau comparatif</div>", unsafe_allow_html=True)
        rows = []
        for m, res in var_results.items():
            row = {"Méthode": m}
            for a in alphas_used:
                r = res.get(a, {})
                row[f"VaR {a*100:.0f}% (€)"]  = f"{r.get('VaR',0):,.0f}"
                row[f"VaR {a*100:.0f}% (%)"]  = f"{r.get('VaR_pct',0)*100:.3f}%"
                row[f"ES {a*100:.0f}% (€)"]   = f"{r.get('ES',0):,.0f}"
            row["Paramètres"] = list(var_results[m].values())[0].get("params","")
            rows.append(row)
        df_var = pd.DataFrame(rows).set_index("Méthode")
        st.dataframe(df_var, use_container_width=True)

        st.markdown("<div class='section-header'>Graphique comparatif</div>", unsafe_allow_html=True)
        fig_v = fig_var_comparaison(var_results, alpha_display, pv)
        st.pyplot(fig_v, use_container_width=True)
        plt.close(fig_v)

        st.markdown("<div class='section-header'>Distribution des rendements</div>", unsafe_allow_html=True)
        fig_d = fig_distribution(port_r, var_results)
        st.pyplot(fig_d, use_container_width=True)
        plt.close(fig_d)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : BACKTESTING
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "🧪 Backtesting":
    st.title("Backtesting des modèles de VaR")

    if st.session_state["var_results"] is None:
        st.info("💡 Calculez d'abord les VaR dans la page **Calcul VaR**.")
        st.stop()

    rendements   = st.session_state["rendements"]
    var_results  = st.session_state["var_results"]
    port_r       = rendements.mean(axis=1).values
    alphas_used  = sorted(list(list(var_results.values())[0].keys()))

    st.markdown("""
    <div class='info-box'>
    <b style='color:#C9A84C'>Test de Kupiec (POF)</b> — Vérifie si la fréquence d'exceptions
    est statistiquement conforme au niveau de confiance déclaré. <b>LR ~ χ²(1).</b>
    <br><br>
    <b style='color:#C9A84C'>Test de Christoffersen (CC)</b> — Teste l'indépendance temporelle
    des exceptions. Un clustering signale un modèle insensible aux chocs de marché.
    <br><br>
    <span style='color:#1db87a'>✅ p-value &gt; 5%</span> → modèle validé &nbsp;&nbsp;
    <span style='color:#e05252'>❌ p-value ≤ 5%</span> → modèle rejeté
    </div>""", unsafe_allow_html=True)

    btn3, _ = st.columns([1,3])
    with btn3:
        bt_btn = st.button("▶  Lancer le backtesting", type="primary", use_container_width=True)

    if bt_btn:
        with st.spinner("Backtesting en cours…"):
            bt_results = {}
            for method, res in var_results.items():
                bt_results[method] = {}
                for a in alphas_used:
                    var_pct = res.get(a, {}).get("VaR_pct", np.nan)
                    if np.isnan(var_pct):
                        continue
                    k  = kupiec_test(port_r, var_pct, a)
                    cc = christoffersen_test(port_r, var_pct)
                    bt_results[method][a] = {"kupiec": k, "cc": cc}
            st.session_state["bt_results"] = bt_results
        st.success("✅ Backtesting terminé.")

    if st.session_state["bt_results"]:
        bt_results = st.session_state["bt_results"]

        st.markdown("<div class='section-header'>Résultats des tests</div>", unsafe_allow_html=True)
        rows = []
        for m, alphas in bt_results.items():
            for a, res in alphas.items():
                k, cc = res["kupiec"], res["cc"]
                rows.append({
                    "Méthode":       m,
                    "CL":            f"{a*100:.0f}%",
                    "Exceptions":    k["N"],
                    "T":             k["T"],
                    "Taux obs.":     f"{k['rate']*100:.2f}%",
                    "Taux att.":     f"{(1-a)*100:.2f}%",
                    "Kupiec LR":     round(k["LR"],3),
                    "Kupiec p":      round(k["p_value"],4),
                    "Kupiec ✓":      "✅ OK" if k["valid"] else "❌",
                    "CC LR_ind":     round(cc.get("LR_ind",0),3),
                    "CC p":          round(cc.get("p_value_ind",0),4),
                    "CC ✓":          "✅ OK" if cc.get("valid") else "❌",
                })
        df_bt = pd.DataFrame(rows)
        st.dataframe(df_bt.set_index("Méthode"), use_container_width=True)

        st.markdown("<div class='section-header'>Graphique p-values (Kupiec)</div>", unsafe_allow_html=True)
        fig_bt = fig_backtesting({m: {a: {"p_value": bt_results[m][a]["kupiec"]["p_value"]}
                                       for a in bt_results[m]} for m in bt_results})
        st.pyplot(fig_bt, use_container_width=True)
        plt.close(fig_bt)

        # Verdict synthétique
        st.markdown("<div class='section-header'>Verdict synthétique</div>", unsafe_allow_html=True)
        alpha_v = alphas_used[-1]
        cols_v = st.columns(len(bt_results))
        for col, (m, alphas) in zip(cols_v, bt_results.items()):
            res = alphas.get(alpha_v, {})
            k, cc = res.get("kupiec",{}), res.get("cc",{})
            k_ok, cc_ok = k.get("valid",False), cc.get("valid",False)
            score = "✅ Validé" if k_ok and cc_ok else ("⚠️ Partiel" if k_ok or cc_ok else "❌ Rejeté")
            color = "#1db87a" if k_ok and cc_ok else ("#C9A84C" if k_ok or cc_ok else "#e05252")
            with col:
                st.markdown(f"""
                <div class='var-card' style='text-align:center'>
                  <div class='var-card-title'>{m}</div>
                  <div style='font-size:15px;font-weight:700;color:{color};margin:6px 0'>{score}</div>
                  <div style='font-size:10px;color:#8aa0bc;font-family:JetBrains Mono,monospace'>
                  Kupiec p={k.get('p_value',0):.4f}<br>CC p={cc.get('p_value_ind',0):.4f}
                  </div>
                </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE : REPORTING
# ══════════════════════════════════════════════════════════════════════════════

elif menu == "📊 Reporting":
    st.title("Génération des Rapports")

    if st.session_state["var_results"] is None:
        st.info("💡 Calculez d'abord les VaR dans la page **Calcul VaR**.")
        st.stop()

    rendements  = st.session_state["rendements"]
    var_results = st.session_state["var_results"]
    bt_results  = st.session_state["bt_results"] or {}
    pv          = st.session_state["pv"] or 10_000_000
    port_r      = rendements.mean(axis=1).values
    alphas_used = sorted(list(list(var_results.values())[0].keys()))

    st.markdown("""
    <div class='info-box'>
    Générez automatiquement un <b>rapport Excel</b> de suivi opérationnel
    et un <b>rapport PDF</b> exécutif destiné à la Direction du Département Risque.
    Les deux fichiers incluent l'ensemble des résultats VaR et de backtesting avec graphiques.
    </div>""", unsafe_allow_html=True)

    st.markdown("<div class='section-header'>Aperçu des résultats</div>", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    alpha_99 = max(alphas_used)
    best = "TVE-GARCH" if "TVE-GARCH" in var_results else list(var_results.keys())[-1]
    var_best = var_results[best][alpha_99]["VaR"]
    es_best  = var_results[best][alpha_99]["ES"]
    n_ok = sum(1 for m in bt_results for a in bt_results[m]
               if bt_results[m][a]["kupiec"]["valid"]) if bt_results else 0
    n_total = sum(len(bt_results[m]) for m in bt_results) if bt_results else 0

    c1.metric("Méthode recommandée", best)
    c2.metric(f"VaR 99% ({best})", f"{var_best/1000:.1f} k€")
    c3.metric(f"ES 99% ({best})", f"{es_best/1000:.1f} k€")
    c4.metric("Backtests validés", f"{n_ok}/{n_total}" if n_total else "—")

    st.markdown("<div class='section-header'>Téléchargements</div>", unsafe_allow_html=True)
    col_xlsx, col_pdf = st.columns(2)

    # Excel
    with col_xlsx:
        st.markdown("""
        <div class='var-card'>
          <div class='var-card-title'>📊 Rapport Excel</div>
          <div style='font-size:12px;color:#d0daea;margin:8px 0;line-height:1.6'>
          3 feuilles : Résumé VaR · Backtesting · Données historiques<br>
          Formatage professionnel, couleurs coded, tableaux structurés.
          </div>
        </div>""", unsafe_allow_html=True)
        if HAS_XLSX:
            with st.spinner("Génération Excel…"):
                bt_fmt = {}
                for m, alphas in bt_results.items():
                    bt_fmt[m] = {}
                    for a, res in alphas.items():
                        bt_fmt[m][a] = res
                xlsx_bytes = generer_excel(rendements, var_results, bt_fmt, pv)
            if xlsx_bytes:
                st.download_button("⬇  Télécharger le rapport Excel",
                                   data=xlsx_bytes,
                                   file_name="VaR_Report.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
        else:
            st.warning("openpyxl non installé : pip install openpyxl")

    # PDF
    with col_pdf:
        st.markdown("""
        <div class='var-card'>
          <div class='var-card-title'>📄 Rapport PDF</div>
          <div style='font-size:12px;color:#d0daea;margin:8px 0;line-height:1.6'>
          5 sections : Couverture · VaR · Distribution · Backtesting · Conclusions<br>
          Graphiques intégrés, mise en page exécutive, prêt à imprimer.
          </div>
        </div>""", unsafe_allow_html=True)
        if HAS_PDF:
            with st.spinner("Génération PDF…"):
                fv  = fig_var_comparaison(var_results, alpha_99, pv)
                fd  = fig_distribution(port_r, var_results)
                pdf_bytes = generer_pdf(var_results, bt_fmt if bt_results else {},
                                        {}, pv, fv, fd)
            if pdf_bytes:
                st.download_button("⬇  Télécharger le rapport PDF",
                                   data=pdf_bytes,
                                   file_name="VaR_Risk_Report.pdf",
                                   mime="application/pdf",
                                   use_container_width=True)
        else:
            st.warning("reportlab non installé : pip install reportlab")

    # Instructions d'installation
    with st.expander("📦 Instructions d'installation"):
        st.code("""# Installer toutes les dépendances
pip install streamlit yfinance pandas numpy scipy \\
            openpyxl reportlab matplotlib

# Lancer l'application
streamlit run app.py""", language="bash")
